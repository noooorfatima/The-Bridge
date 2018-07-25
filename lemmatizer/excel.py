"""
Utility functions for bridge-tools scripts pertaining to writing data to
Excel spreadsheets using openpyxl.
"""


from os.path import normpath, splitext, commonprefix, basename

# openpyxl.readthedocs.io
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.writer.write_only import WriteOnlyCell
from openpyxl.styles import Font

from column import *

def openSpreadsheet(path):
    """
    Returns an openpyxl.Workbook object loaded from the .xlsx file at `path`
    if it exists, or creates a new WriteOnlyWorksheet if it does not.

    Parameters:
        path (str): the address of the file

    Returns:
        an openpyxl workbook object
    """
    try:
        workbook = load_workbook(path)
    except IOError:
        workbook = Workbook(write_only=True)
    return workbook

def saveGroupsToSpreadsheets(groups, columns, *, path=None, file_prefix=None,
                             append=False, sheet_title="Sheet1", echo=False):
    """
    Create a new output spreadsheet for each key : subiterable pair in
    `groups` whose sheet contains values associated with the values of the
    subiterable. Save these spreadsheets to the directory `path` with
    filenames given by `file_prefix` and the keys of `groups`.

    Refer to createWorkbooksFromGroups() for more information.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        path (str): the path to the directory in which to save files
        file_prefix (str): a title prepended onto keys when naming files
        append (Optional[bool]): append data to an existing file if found;
            note that any sheets created from the same groupings as `groups`
            and with the same `sheet_prefix` WILL be overwritten
        sheet_title (str): the title of the active worksheet of each workbook
    """
    filename = '{}-{{}}'.format(file_prefix)
    for key, group in groups:
        saveDataToSpreadsheet(group, columns, filename.format(key), echo=echo,
                              path=path, sheet_title=sheet_title, append=append)

def saveGroupsToSpreadsheet(groups, columns, filename, *, path=None,
                            append=False, sheet_prefix=None, echo=False):
    """
    Create a new output spreadsheet with a sheet for each key : subiterable
    pair in `groups` containing values associated with the values of the
    subiterable. Save this spreadsheet to the directory `path` with the
    filename `filename`.

    Refer to createWorkbooksFromGroups() for more information.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        path (str): the path to the directory in which to save the file
        filename (str): the filename to use when saving spreadsheet (excludes
            extension, which is always .xlsx)
        append (Optional[bool]): append data to an existing file if found;
            note that any sheets created from the same groupings as `groups`
            and with the same `sheet_prefix` WILL be overwritten
        sheet_prefix (Optional[str]): a title prepended onto keys when naming
            worksheets
    """
    destination = '{path}{filename}.xlsx'.format(
        path = (normpath(path) + '/') if path else '', filename = filename
    )

    if append:
        workbook = load_workbook(destination)
        createWorksheetsFromGroups(groups, columns, workbook,
                                   sheet_prefix=sheet_prefix)
    else:
        workbook = createWorkbookFromGroups(groups, columns,
                                            sheet_prefix=sheet_prefix)

    print("Saving changes to {}".format(destination))
    workbook.save(destination)

def saveDataToSpreadsheet(data, columns, filename, *, path=None, append=False,
                          sheet_title="Sheet1", echo=False):
    """
    Create a new output spreadsheet whose sheet contains values associated
    with the values of `data`. Save this spreadsheet to the directory `path`
    with the filename `filename`.

    Refer to createWorkbookFromData() for more information.

    Parameters:
        data (Iterable): an iterable over values to be passed as arguments to
            the `find_value` coordinates of the Column tuples in `columns`
        columns (Iterable[Column]): an iterable over Column tuples
        path (str): the path to the directory in which to save the file
        filename (str): the filename to use when saving spreadsheet (excludes
            extension, which is always .xlsx)
        append (Optional[bool]): append data to an existing file if found
        sheet_title (str): the title of the active worksheets of each workbook
    """
    destination = '{path}{filename}.xlsx'.format(
        path = (normpath(path) + '/') if path else '', filename = filename
    )

    if append:
        workbook = load_workbook(destination)
        try:
            appendDataToWorksheet(workbook[sheet_title], data, columns)
        except KeyError:
            createWorksheetFromData(data, columns, workbook,
                                    sheet_title=sheet_title)
    else:
        workbook = createWorkbookFromData(data, columns,
                                          sheet_title=sheet_title)
    workbook.save(destination)
    print("Saving changes to {}".format(destination))

def createWorkbooksFromGroups(groups, columns, *, sheet_title="Sheet1"):
    """
    Create a new output spreadsheet for each key : subiterable pair in
    `groups` whose sheet contains values associated with the values of the
    subiterable. Return these spreadsheets as a list of key : Workbook tuples.

    Refer to createWorkbookFromData() for more information.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        sheet_title (str): the title of the active worksheets of each workbook

    Returns:
        a list of key (str) : workbook (openpyxl.Workbook) pairs
    """
    workbooks = []
    for key, group in groups:
        workbook = createWorkbookFromData(group, columns, sheet_title=sheet_title)
        workbooks.append((key, workbook))
    return workbooks

def createWorkbookFromData(data, columns, *, sheet_title="Sheet1"):
    """
    Create a new output spreadsheet whose sheet contains values associated
    with the values of `data`.

    Refer to createWorksheetFromData() for more detailed information.

    Parameters:
        data (Iterable): an iterable over values to be passed as arguments to
            the `find_value` coordinates of the Column tuples in `columns`
        columns (Iterable[Column]): an iterable over Column tuples
        sheet_title (str): the title of the active worksheet of `workbook`

    Returns:
        an openpyxl workbook object
    """
    workbook = Workbook(write_only = True)
    createWorksheetFromData(data, columns, workbook, title=sheet_title)
    return workbook

def createWorkbookFromGroups(groups, columns, *, sheet_prefix=None):
    """
    Create a new output spreadsheet whose sheets contains values associated
    with the values of the subiterable in each group of `groups`.

    Refer to createWorksheetsFromGroups() for more detailed information.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        sheet_prefix (Optional[str]): a title prepended onto keys when naming
            worksheets

    Returns:
        an openpyxl workbook object
    """
    workbook = Workbook(write_only=True)
    createWorksheetsFromGroups(groups, workbook, columns,
                               sheet_prefix=sheet_prefix)
    return workbook

def createWorksheetsFromGroups(groups, columns, workbook, *,
                               sheet_prefix=None):
    """
    Create new output worksheets in the spreadsheet `workbook` from the
    key-subiterable pairs in `groups`.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        workbook (openpyxl.Workbook): the workbook in which to create the sheet
        sheet_prefix (str): a title prepended onto keys when titling worksheets
    """
    for key, group in groups:
        if key is None and sheet_prefix is None:
            createWorksheetFromData(group, workbook, columns)
            continue
        elif key is not None and sheet_prefix is not None:
            title = "{} {}".format(sheet_prefix, key)
        else:
            title = str(key) if key is not None else str(sheet_prefix)
        createWorksheetFromData(group, workbook, columns, title=title)

def createWorksheetFromData(data, columns, workbook,title="Sheet1",
                            use_bold_headers=True):
    """
    Create a new output worksheet in the spreadsheet `workbook` and append
    values associated with `data` into the columns represented by `columns`.

    The values of the first row of the worksheet are obtained by the `name`
    coordinates of the Column tuples in `columns`.

    Otherwise, the value for a cell in row i and column j is obtained by
    passing the 'i - 1'th element of `data` and the integer i as arguments
    to the  `find_value` coordinate of the Column tuple in `columns` whose
    `number` coordinate is j. If no such Column exists, the cell will be empty.

    Parameters:
        data (Iterable): an iterable over values to be passed as arguments to
            the `find_value` coordinates of the Column tuples in `columns`
        columns (Iterable[Column]): an iterable over Column tuples
        workbook (openpyxl.Workbook): the workbook in which to create the sheet
        title (str): the title of the worksheet to be created
    """
    # Create the workbook and worksheet:
    worksheet = workbook.create_sheet(title=title)
    appendColumnHeadersToWorksheet(columns, worksheet, use_bold=use_bold_headers)
    appendDataToWorksheet(worksheet, data, columns)

def appendColumnHeadersToWorksheet(columns, worksheet, *, use_bold=True):
    """
    Appends the names of the Column objects in `columns` onto the excel
    worksheet `worksheet` as headers.

    Parameters:
        columns (Iterable[Column]): an iterable over Column tuples
        worksheet (openpyxl.worksheet.worksheet.Worksheet):
            the worksheet onto which headers will be appended
    """
    headers = [None] * max(column.number for column in columns)
    for column in columns:
        cell = WriteOnlyCell(worksheet, value = column.name)
        if use_bold:
            cell.font = Font(bold = True)
        headers[column.number - 1] = cell
    worksheet.append(headers)

def appendDataToWorksheet(worksheet, data, columns, *, verbose=True):
    """
    Appends values associated with the iterable `data` onto the excel worksheet
    `worksheet`.

    Parameters:
        data (Iterable): an iterable over values to be passed as arguments to
            the `find_value` coordinates of the Column tuples in `columns`
        columns (Iterable[Column]): an iterable over Column tuples
        worksheet (openpyxl.worksheet.worksheet.Worksheet):
            the worksheet onto which data will be appended
        verbose (Optional[bool]): whether to print a message on completion
    """
    total_rows = 0
    for i, word in enumerate(data, start=worksheet._max_row):
        row = [None] * max(column.number for column in columns)
        for column in columns:
            row[column.number - 1] = column.find_value(word, i)
        worksheet.append(row)
        total_rows += 1
    if verbose: print("Wrote {} rows to {}".format(total_rows, worksheet.title))
