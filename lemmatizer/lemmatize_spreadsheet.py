import openpyxl
from openpyxl import load_workbook #for some reason this insisted on being imported seperately. If this line can be deleted later, please do
from .lemma import LemmaReplacer
from cltk.stem.latin.j_v import JVReplacer
from cltk.tokenize.word import WordTokenizer, nltk_tokenize_words
from .autoLemma import processUnicodeDecomposition, removeMacrons, removeDiareses, changeGraveAccents
"""
NOTE: THIS MUST RUN IN PYTHON3 BECAUSE IT USES THE CLTK
http://docs.cltk.org/en/latest/latin.html#lemmatization
Should take a langauge and a file as input. you may also specify which sheet in the xlsx file you are using, and which colummn must be lemmatized
"""
def lemmatize(language, filename):
    print( "opening spreadsheet to lemma")
    print((filename,  "filename"))
    print((langauge, 'language'))
    if language == 'latin':
        lemmatize_latin(filename)
    elif language == 'greek':
        lemmatize_greek(filename)

def lemmatize_latin(file, sheet= 'Sheet1', column_to_edit =1): #file should be a .xlsx
    print ("lemmatizing latin")
    print(( "filename", filename))
    tokenizer = WordTokenizer('latin')
    wb = load_workbook(file)
    lemmatizer = LemmaReplacer('latin')
    worksheet = wb[sheet]
    for col_cells in worksheet.iter_cols(min_col = column_to_edit, max_col = column_to_edit):
        for cell in col_cells:
            cell.value = str(cell.value)
       	    print((cell.value, 'value as a string'))
       	    cell.value = cell.value.lower()
            cell.value = jv_replacer.replace(cell.value)
            cell.value = processUnicodeDecomposition(cell.value, removeMacrons)
            cell.value = tokenizer.tokenize(cell.value)
            print((cell.value, 'lowercased, macronless, jv-replaced, tokenized, value', 'woah english participles suck'))
            cell.value = lemmatizer.lemmatize(cell.value, return_string=True)
            print((cell.value, 'lemmatized value'))
            
        
    wb.save(str(filename) + '.xlsx')
	 

def lemmatize_greek(file, sheet = 'Sheet1', column_to_edit = 1): #file should be a .xlsx
    print ("lemmatizing greek")
    lemmatizer = LemmaReplacer('greek')
    wb = load_workbook(file)
    worksheet = wb[sheet]
    for col_cells in worksheet.iter_cols(min_col = column_to_edit, max_col = column_to_edit):
        for cell in col_cells:
            cell.value = str(cell.value)
            print((cell.value, 'value as a string'))
            cell.value = cell.value.lower()
            cell.value = processUnicodeDecomposition(cell.value, removeDiareses, changeGraveAccents)
            cell.value = tokenizer.tokenize(cell.value)
            print((cell.value, 'lowercased, and other weird greek things removed, and tokenized value'))
            cell.value = lemmatizer.lemmatize(cell.value, return_string=True)
            print((cell.value, 'lemmatized value'))

    wb.save(str(filename) + '.xlsx')



if __name__ == "__main__":
    import sys
    print((sys.argv))
    lemmatize(sys.argv[1], sys.argv[2])
