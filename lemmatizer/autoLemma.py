
#!/usr/bin/env python
# coding=utf-8

"""
Identifies possible lemmata and snopyses for a set of Latin or Greek words.

Uses the lemmatizer provided by the Classical Languages Toolkit.
Must take a plain text file (.txt) as input; as output, a new spreadsheet
is generated showing the input words and matched info about them.

Locations should be separated from the rest of the text by square brackets.
Examples: [1], [2.2], [3.1.2], ... If there are no location markers, then
line numbers will be used.

Note that docopt, openpyxl, regex, and the Classical Languages Toolkit must be
installed as dependencies, as well as the relevant corpora: refer to
http://docs.cltk.org/en/latest/importing_corpora.html#importing-a-corpus.

Usage:
    autoLemma.py (greek|latin) [options] [--] <file> ...

Arguments:
    language                     only Greek and Latin are currently supported
    <file>                       one or more plain text files to be lemmatized

Options:
    -h --help                    show this help message and exit
    -a, --include-ambiguous      choose most likely lemma for an ambiguous form
    -f <format>, --l <format>    format of lemmata [default: cltk]       # TODO
    -o <file>, --output <file>   output file (uses input filename by default)
    --append                     append onto an existing output file
    --dir <path>                 prefixed onto file addresses for input/output
    -e, --echo                   print values of output spreadsheets to console
    --force-vi                   force lemmata to use "V, I" in place of "U, J"
    --force-ui                   force lemmata to use "U, I" in place of "V, J"
    --force-lowercase-lemmata    force lemmata to be lowercase
    --force-no-trailing-digits   force lemmata to not contain trailing digits
    --force-no-punctuation       force lemmata to not contain punctuation
    --force-uppercase-lemmata    force lemmata to be uppercase
    --group-by (section|location|file)  split-into groupings [default: section]
    --formulae                   also populate cells with formulae
    --split-into (files|sheets)  split output by (section) into sheets or files
    --text-name <title>          text name (for multiple input or output files)
    --output-sheet <name>        sheet name for output [default: LEMMATA MATCH]
    --use-line-numbers           include line numbers in word locations
    --use-sections               include column with simple section numbers
    --use-detailed-sections      include column with full location w/o line #s
"""

from sys import exit
from os.path import normpath, splitext, commonprefix, basename
from mimetypes import guess_type
from collections import namedtuple
from itertools import groupby
from unicodedata import normalize
from warnings import warn
from string import digits
import codecs


# https://pypi.python.org/pypi/regex
# this is to recognize accented Greek characters as alphabetic
#import regex
import re as regex



# http://docs.cltk.org/en/latest/latin.html#lemmatization
from lemma import LemmaReplacer
from cltk.stem.latin.j_v import JVReplacer
from cltk.tokenize.word import WordTokenizer, nltk_tokenize_words

# https://github.com/docopt/docopt
from docopt import docopt
#everything above here works
# openpyxl.readthedocs.io
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.write_only import WriteOnlyCell
from column import *
from openpyxl.utils import get_column_letter





# TODO these should be patched in CLTK
IGNORED_LEMMATA = ['publica', 'tanto', 'multi', 'verro', 'medio', 'privo',
                   'consento', 'quieto', 'mirabile', 'retineo', 'subeo',
                   'Arruntius', 'disparo', 'prius', 'scelerato']

Word = namedtuple('Word', ['form', 'lemma', 'location'])
Word.__doc__ = \
    """
    A `Word` is a tuple with `form`, `lemma`, and `location` coordinates.

    Coordinates:
        form (str): the inflected word
        lemma (str or NoneType): the lemma which corresponds to `form`
        location (str): a string indicating the location of the word within
            a certain text; of the form 1, 2, 3, ... or 1.1, 1.2, 2.1, ...
    """

Location = namedtuple('Location', ['label', 'text'])
Location.__doc__ = \
    """
    A `Location` is a tuple with `label` and `text` coordinates.

    Coordinates:
        label (str): the label for the location; of the form 1, 2, 3, ...,
            or 1.1, 1.2, 2.1, ...
        text (str): the text contained in this location
    """

Column = Column
Column.__doc__ += \
    """
    In this script, `find_value` coordinates of Column tuples will be passed
    Word tuples for their first parameter.
    """

CHECK_FORMULA = "=VLOOKUP({}{},'LEMMATA-Vocab'!A:A,1,FALSE)"
# COUNT_FORMULA = "=ROW({}{}) - 1"
DISPLAY_FORMULA = "=VLOOKUP({}{},'LEMMATA-Vocab'!C:E,2,FALSE)"
SHORTDEF_FORMULA = "=VLOOKUP({}{},'LEMMATA-Vocab'!C:E,3,FALSE)"
LONGDEF_FORMULA = "=VLOOKUP({}{},'LEMMATA-Vocab'!C:F,4,FALSE)"
LOCALDEF_FORMULA = "=VLOOKUP({}{},'LEMMATA-Vocab'!C:G,5,FALSE)"

FORMULAE_COLUMNS = [
    "CHECK", "DISPLAY LEMMA", "SHORTDEF", "LONGDEF", "LOCALDEF"
]

OUTPUT_COLUMNS = [
    Column("CHECK", 1,
           lambda __, row: CHECK_FORMULA.format(LEMMA_COLUMN_LETTER, row)),
    Column("TITLE", 2,
           lambda word, __: word.lemma if word.lemma is not None else ''),
    Column("TEXT", 3,
           lambda word, __: word.form),
    Column("LOCATION", 4,
           lambda word, __: word.location),
    Column("SECTION", 5,
           lambda word, __: word.location),
    Column("RUNNING COUNT", 6,
           lambda __, row: str(row)),
    Column("DISPLAY LEMMA", 7,
           lambda __, row: DISPLAY_FORMULA.format(LEMMA_COLUMN_LETTER, row)),
    Column("SHORTDEF", 8,
           lambda __, row: SHORTDEF_FORMULA.format(LEMMA_COLUMN_LETTER, row)),
    Column("LONGDEF", 9,
           lambda __, row: LONGDEF_FORMULA.format(LEMMA_COLUMN_LETTER, row)),
    Column("LOCALDEF", 10,
           lambda __, row: LOCALDEF_FORMULA.format(LEMMA_COLUMN_LETTER, row))
]

OUTPUT_COLUMNS_WITHOUT_FORMULAE = [
    column for column in OUTPUT_COLUMNS if column.name not in FORMULAE_COLUMNS
]

LEMMA_COLUMN_LETTER = get_column_letter(
    getColumnByName(OUTPUT_COLUMNS, "TITLE").number
)

class NLTKTokenizer(WordTokenizer):
    """
    A wrapper for the `nltk_tokenize_words` function that inherits from the
    cltk.tokenize.word.WordTokenizer class. This could be expanded to implement
    some special tokenization for Greek if needed.
    """
    def __init__(self, language):
        self.language = language
    def tokenize(self, string):
        return nltk_tokenize_words(string)

def processUnicodeDecomposition(string, *functions):
    """
    Apply all functions in arguments to the canonicial decomposition of the
    unicode string `string` (i.e. with combining characters such as accents
    separated from the characters they combine with).

    Parameters:
        string (str): a Unicode string
        *functions (List[Callable]): variable length list of functions

    Returns:
        (str): the canonical normalization of `string` after each function in
            `*functions` has been applied to its decomposition
    """
    decomposition = normalize('NFD', string)
    for function in functions:
        decomposition = function(decomposition)
    return normalize('NFC', decomposition)

def removeMacrons(string):
    """"
    Removes the macrons from any macron-ed characters in a string.

    Parameters:
        string (str): the string whose macrons are to be removed
            macrons must be separated as combining characters

    Returns:
        (str): `string` without any macrons
    """
    return regex.sub('\\u0304', '', string)

def removeDiareses(string):
    """"
    Removes any diareses in string.

    Parameters:
        string (str): the string whose diareses are to be removed
            diareses must be separated as combining characters

    Returns:
        (str): `string` without any diareses
    """
    return regex.sub('\\u0308', '', string)

def changeGraveAccents(string):
    """
    Changes any grave accents in `string` to acute accents.

    Parameters:
        string (str): the string whose grave accents are to be replaced
            grave accents must be separated as combining characters

    Returns:
        (str): `string` with acute accents in place of graves
    """
    str = regex.split('\\u0300', string)
    out = ''
    for x in range(0,len(str),1):
        #y = str[x+1] if len(str) > x+1 else  ''
        out += str[x]  + u'\u0301' #+ y
    #if not len(str)%2==0:
        #out+=str[len(str)-1]+ u'\u0301' 
    return out
    #print(regex.sub('\\u0300', '\\u0301', string))
   # return regex.sub('\\u0300', '\\u0301', string)

def sectionFromWord(word):
    """
    Returns the section associated with the location coordinate of the Word
    tuple `word`. The section is given by the first identifiable group of
    characters in a location split by delimeters or mixed between alphabetic
    and numeric characters.

    For example, the locations '1.1' and '1.2' are both considered to belong to
    section '1'; likewise, '9a' and '9b' both belong to section '9'.

    Parameters:
        word (Word): the Word tuple whose section is to be found

    Returns:
        section (str): the section associated with the location of `word`
    """
    match = regex.match(r'(?i)[0-9]+|[A-Z]+|[Α-Ω]+',word.location)
    return match.group() if match is not None else None

def detailedSectionFromWord(word):
    """
    Returns the detailed section associated with the location coordinate of
    the Word tuple `word`. The detailed section is given by all but the last
    identifiable groups of characters in a location split by delimeters or
    mixed between alphabetic and numeric characters.

    For example, the locations '1.1.3' and '1.2.4' are both considered to
    belong to sections '1.1' and '1.2' respecively; likewise, '9a1' and '9b1'
    belong to sections '9a' and '9b'.

    Parameters:
        word (Word): the Word tuple whose section is to be found

    Returns:
        section (str): the detailed section associated with the location
            coordinate of `word`
    """
    match = regex.match(r'(?i)[0-9]+|[A-Z]+|[Α-Ω]+', word.location[::-1])
    if match is not None:
        return word.location[:-1 * match.end()].rstrip('.')

def groupWordsBySection(words, detailed=False):
    """
    Groups the Word tuples of `words` by the sections associated with their
    location coordinates as identified by `sectionFromWord` or by
    `detailedSectionFromWord`.

    Parameters:
        words (List[Word]): an iterable over Word tuples sorted by location
        detailed (bool): whether to use the `detailedSectionFromWord` function

    Returns:
        an iterator over tuples of section (str): subiterable over `words`
    """
    return groupby(words, sectionFromWord)

def groupWordsByLocation(words):
    """
    Groups the Word tuples of `words` by unique location coordinates.

    Parameters:
        words (List[Word]): an iterable over Word tuples sorted by location

    Returns:
        an iterator over tuples of location (str): subiterable over `words`
    """
    return groupby(words, lambda word: word.location)

def lemmatizeToken(token, lemmatizer):
    """
    Returns the lemma corresponding to the token `token`.

    `Token` should be already formatted for CLTK lemmatization (i.e.
    tokenized into a single word, with JV-Replacement if Latin,
    in Unicode format for Greek).

    Parameters:
        string (string): the string being lemmatized
        lemmatizer (cltk.stem.lemma.LemmaReplacer): the lemmatizer to use to
            identify the lemma of `token`

    Returns:
        lemma (string): the lemma which corresponds to `token`,
            or `None` if no such lemma was found

    Raises:
        ValueError if `token` contains non-alphabetic characters,
            multiple lemmata, or only whitespace
    """

    token = token.lower()
#see below
    if not token=='́':
        lemmata = lemmatizer.lemmatize(token) #, default = '')
    else:
        return None

    if len(lemmata) > 1:
        raise ValueError("\'{}\' contains multiple lemmata".format(token))
    elif len(lemmata) == 0:
        raise ValueError("\'{}\' is empty or nonalphabetic.".format(token))

   # if token != '́' and not token.isspace():
    lemma = lemmata[0]
    if not lemma:
        # To prevent problems with case (which should be fixed in CLTK with the lemmatizer rewrite)
        lemma = lemmatizer.lemmatize(token.capitalize(), default = '')[0]
    if lemma in IGNORED_LEMMATA:
        lemma = ""         # For some known bugs with CLTK

    return lemma if lemma else None
    #else:
        #return None

def locationsFromFile(file, *, use_line_numbers = False):
    """
    Iterates over unique locations as indicated by section delimiters or line
    numbers in the text file `file`.

    Parameters:
        file (io.TextIOWrapper): the file being read
        use_line_numbers (bool): whether to append line numbers to locations
            when section delimiters are present

    Yields:
        location (Location): a Location tuple of the next unique location in
            `file` and the text it contains
    """
    line_number = 1
    text, section = '', ''

    def getFormattedLabel():
        if section and use_line_numbers:
            return '{}.{}'.format(section, line_number)
        else:
            return section if section else str(line_number)

    for line_text in file:
        # split by lines
        if regex.search(r'[0-9]+$', line_text):
            # found inline line number
            line_number = int(regex.search(r'[0-9]+$', line_text).group())
        for string in regex.split(r'(\[[0-9.]+\])', line_text):
            # split by delimeters
            # http://stackoverflow.com/questions/2136556/in-python-how-do-i-split-a-string-and-keep-the-separators
            if regex.search(r'\[[0-9.]+\]', string):
                # found section delimiter
                if text and section and not text.isspace():
                    yield Location(getFormattedLabel(), text)
                # write section delimeter to location
                text, section, line_number = '', string[1:-1], 1
            elif string and not string.isspace():
                text += string
        if text and (use_line_numbers or not section):
            yield Location(getFormattedLabel(), text)
            text, line_number = '', line_number + 1

    yield Location(getFormattedLabel(), text)

def wordsFromFile(file, lemmatizer, *, use_line_numbers = False):
    """
    Extracts words as Word tuples from the text file `file`.

    Parameters:
        file (io.TextIOWrapper): the file being read
        lemmatizer (cltk.stem.lemma.LemmaReplacer): the lemmatizer to use to
            identify lemmata of words in `file`

    Yields:
        word (Word): a Word tuple of the next form to appear in `file`
    """
    jv_replacer = JVReplacer()
    try:
        tokenizer = WordTokenizer(lemmatizer.language)
    except AssertionError:
        # Language of `lemmatizer` does not support CLTK tokenization
        # Clitics must be separated by spacing
        tokenizer = NLTKTokenizer(lemmatizer.language)

    locations = locationsFromFile(file, use_line_numbers=use_line_numbers)
    for location, text in locations:
        if lemmatizer.language == 'latin':
            text = jv_replacer.replace(text)
            text = processUnicodeDecomposition(text, removeMacrons)
        elif lemmatizer.language == 'greek':
            text = processUnicodeDecomposition(text, removeDiareses,
                                               changeGraveAccents)
        for token in tokenizer.tokenize(text):
            for form in regex.split(r'(?:\P{L},.+)', token):
                # split around punctuation
                try:
                    if form == ',' or form == '.' or form == ':' or form == '(' or form == ')' or form == ';' or form == ']' or form == '[' or form == '?':
                        pass
                    else:
                        if not token=='́':
                            lemma = lemmatizeToken(form, lemmatizer)
                            yield Word(form, lemma, location)
                        else: 
                            continue
                except ValueError:
                    # token contains non-alphabetic characters which
                    # aren't leading or trailing, or contains no
                    # alphabetic characters
                    continue

def wordsFromPathList(paths, lemmatizer, **kwargs):
    """
    Extracts words as Word tuples from the files given by `paths`. Note:
    removes paths from `paths` successively during iteration.

    Parameters:
        paths (List[str]): each path must be the address of a plain text file
        lemmatizer (cltk.stem.lemma.LemmaReplacer): the lemmatizer to use to
            identify lemmata of words in `file`

    Yields:
        word (Word): the next Word tuple to appear in the files in `paths`
    """
    while len(paths) > 0:
        path = paths.pop(0)
        file_type = guess_type(path)[0]
        if file_type != 'text/plain':
            warn("Only plain text input is supported: "
                 "{} appears to be {}.".format(path, file_type))
        print(("Loading {}".format(path)))
        try:
            with codecs.open(path, 'r', 'utf-8') as file:
                words = wordsFromFile(file, lemmatizer, **kwargs)
                yield from words
        except IOError:
            exit("Could not find file in path {}".format(path))

def autoLemma(args, *, lemmatizer=None, wordsFromPathList=wordsFromPathList):
    """
    Generates lemmatized spreadsheets from command-line arguments given by
    `args` (see docstring of `autoLemma.py`) using the function
    `wordsFromPathList` to extract Word tuples from the paths in
    `args['<file>']`.

    Parameters:
        args (dict of str:str): see docstring of autoLemma.py
        lemmatizer (cltk.stem.lemma.LemmaReplacer): the lemmatizer to use to
            identify lemmata of words
        wordsFromPathList (Callable): returns an iterable over Word tuples from
            file paths
    """
    print("I AM NOT DEAD YET")
    if lemmatizer is None:
        lemmatizer = LemmaReplacer('latin' if args['latin'] else 'greek', include_ambiguous=args['--include-ambiguous'])
    print('set lemmatizer')
    # Find name of text
    if args['--output']:
        text_name, text_extension = splitext(args['--output'])
        if text_extension != 'xlsx':
            warn("autoLemma.py only outputs .xlsx files")
    elif args['--text-name']:
        text_name = args['--text-name']
    else:
        text_name = splitext(basename(commonprefix(args['<file>'])))[0]
        if not text_name: text_name = 'output'
    # Create input/output path
    path_prefix = (normpath(args['--dir']) + '/') if args['--dir'] else ''

    # Extract Word tuple iterators from input
    paths = [path_prefix+path for path in args['<file>']]
    try:
        words = wordsFromPathList(paths,
                                  lemmatizer=lemmatizer,
                                  use_line_numbers=args['--use-line-numbers'])
    except IOError:
        exit("Error opening one or more files! {!s}".format(args['file']))

    # Group words if necessary by specified means
    if args['--split-into']:
        if args['--group-by'] == 'file':
            groups = groupby(words, lambda __: len(args['<file>']) - len(paths))
        elif args['--split-into'] and args['--group-by'] == 'location':
            groups = groupWordsByLocation(words)
        elif args['--split-into'] and args['--group-by'] == 'section':
            if args['--use-detailed-sections']:
                groups = groupWordsBySection(words, detailed=True)
            else:
                groups = groupWordsBySection(words)
        else:
            exit("Invalid --group-by argument")

    # Write words and matched lemmata to new spreadsheets
    if args['--formulae']:
        columns = OUTPUT_COLUMNS
    else:
        columns = OUTPUT_COLUMNS_WITHOUT_FORMULAE

    if args['--use-detailed-sections']:
        replaceColumnFunction(columns, "SECTION",
                                    lambda word, __: detailedSectionFromWord(word))
    elif not args['--use-sections']:
        columns = [ column for column in columns if column.name != "SECTION" ]

    if args['--force-lowercase-lemmata']:
        wrapColumnFunction(columns, "TITLE", lambda lemma: lemma.lower())

    if args['--force-uppercase-lemmata']:
        wrapColumnFunction(columns, "TITLE", lambda lemma: lemma.upper())

    if args['--force-no-trailing-digits']:
        wrapColumnFunction(columns, "TITLE",
                                 lambda lemma: lemma.rstrip(digits))

    if args['--force-no-trailing-digits']:
        wrapColumnFunction(columns, "TITLE",
                                 lambda lemma: regex.sub(r'\P{L}', '', lemma))

    if args['--force-vi'] or args['--force-ui']:
        if args['--force-vi']:
            replacements = [('u', 'v'), ('U', 'V'), ('j', 'i'), ('J', 'I')]
        elif args['--force-ui']:
            replacements = [('v', 'u'), ('V', 'U'), ('j', 'i'), ('J', 'I')]
        def replace(lemma):
            for pattern, repl in replacements:
                lemma = regex.sub(pattern, repl, lemma)
            return lemma
        wrapColumnFunction(columns, "TITLE", lambda lemma: replace(lemma))

    if args['--split-into'] == 'files':
        saveGroupsToSpreadsheets(
            groups, columns, path=path_prefix, file_prefix=text_name,
            append=args['--append'], sheet_title=args['--output-sheet'],
            echo=args['--echo']
        )
    elif args['--split-into'] == 'sheets':
        saveGroupsToSpreadsheet(
            groups, columns, text_name, path=path_prefix, echo=args['--echo'],
            append=args['--append'], sheet_prefix=args['--output-sheet']
        )
    else:
        saveDataToSpreadsheet(
            words, columns, text_name, path='/tmp', echo=args['--echo'],
            append=args['--append'], sheet_title=args['--output-sheet'],
        )

'''
Formerly excel.py, but it was having problems importing openpyxl
'''
"""
Utility functions for bridge-tools scripts pertaining to writing data to
Excel spreadsheets using openpyxl.
"""


from os.path import normpath, splitext, commonprefix, basename


def openSpreadsheet(path):
    """
    Returns an openpyxl.Workbook object loaded from the .xlsx file at `path`
    if it exists, or creates a new WriteOnlyWorksheet if it does not.

    Parameters:
        path (str): the address of the file

    Returns:
        an openpyxl workbook object
    """
    try:
        workbook = load_workbook(path)
    except IOError:
        workbook = Workbook(write_only=True)
        ws = wb.create_sheet()

    return workbook

def saveGroupsToSpreadsheets(groups, columns, *, path=None, file_prefix=None,
                             append=False, sheet_title="Sheet1", echo=False):
    """
    Create a new output spreadsheet for each key : subiterable pair in
    `groups` whose sheet contains values associated with the values of the
    subiterable. Save these spreadsheets to the directory `path` with
    filenames given by `file_prefix` and the keys of `groups`.

    Refer to createWorkbooksFromGroups() for more information.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        path (str): the path to the directory in which to save files
        file_prefix (str): a title prepended onto keys when naming files
        append (Optional[bool]): append data to an existing file if found;
            note that any sheets created from the same groupings as `groups`
            and with the same `sheet_prefix` WILL be overwritten
        sheet_title (str): the title of the active worksheet of each workbook
    """
    filename = '{}-{{}}'.format(file_prefix)
    for key, group in groups:
        saveDataToSpreadsheet(group, columns, filename.format(key), echo=echo,
                              path=path, sheet_title=sheet_title, append=append)

def saveGroupsToSpreadsheet(groups, columns, filename, *, path=None,
                            append=False, sheet_prefix=None, echo=False):
    """
    Create a new output spreadsheet with a sheet for each key : subiterable
    pair in `groups` containing values associated with the values of the
    subiterable. Save this spreadsheet to the directory `path` with the
    filename `filename`.

    Refer to createWorkbooksFromGroups() for more information.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        path (str): the path to the directory in which to save the file
        filename (str): the filename to use when saving spreadsheet (excludes
            extension, which is always .xlsx)
        append (Optional[bool]): append data to an existing file if found;
            note that any sheets created from the same groupings as `groups`
            and with the same `sheet_prefix` WILL be overwritten
        sheet_prefix (Optional[str]): a title prepended onto keys when naming
            worksheets
    """
    destination = '{path}{filename}.xlsx'.format(
        path = (normpath(path) + '/') if path else '', filename = filename
    )

    if append:
        workbook = load_workbook(destination)
        createWorksheetsFromGroups(groups, columns, workbook,
                                   sheet_prefix=sheet_prefix)
    else:
        workbook = createWorkbookFromGroups(groups, columns,
                                            sheet_prefix=sheet_prefix)

    print("Saving changes to {}".format(destination))
    workbook.save(destination)

def saveDataToSpreadsheet(data, columns, filename, *, path=None, append=False,
                          sheet_title="Sheet1", echo=False):
    """
    Create a new output spreadsheet whose sheet contains values associated
    with the values of `data`. Save this spreadsheet to the directory `path`
    with the filename `filename`.

    Refer to createWorkbookFromData() for more information.

    Parameters:
        data (Iterable): an iterable over values to be passed as arguments to
            the `find_value` coordinates of the Column tuples in `columns`
        columns (Iterable[Column]): an iterable over Column tuples
        path (str): the path to the directory in which to save the file
        filename (str): the filename to use when saving spreadsheet (excludes
            extension, which is always .xlsx)
        append (Optional[bool]): append data to an existing file if found
        sheet_title (str): the title of the active worksheets of each workbook
    """
    destination = '{path}{filename}.xlsx'.format(
        path = (normpath(path) + '/') if path else '', filename = filename
    )

    if append:
        workbook = load_workbook(destination)
        try:
            appendDataToWorksheet(workbook[sheet_title], data, columns)
        except KeyError:
            createWorksheetFromData(data, columns, workbook,
                                    sheet_title=sheet_title)
    else:
        workbook = createWorkbookFromData(data, columns,
                                          sheet_title=sheet_title)
    workbook.save(destination)
    print("Saving changes to {}".format(destination))

def createWorkbooksFromGroups(groups, columns, *, sheet_title="Sheet1"):
    """
    Create a new output spreadsheet for each key : subiterable pair in
    `groups` whose sheet contains values associated with the values of the
    subiterable. Return these spreadsheets as a list of key : Workbook tuples.

    Refer to createWorkbookFromData() for more information.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        sheet_title (str): the title of the active worksheets of each workbook

    Returns:
        a list of key (str) : workbook (openpyxl.Workbook) pairs
    """
    workbooks = []
    for key, group in groups:
        workbook = createWorkbookFromData(group, columns, sheet_title=sheet_title)
        workbooks.append((key, workbook))
    return workbooks

def createWorkbookFromData(data, columns, *, sheet_title="Sheet1"):
    """
    Create a new output spreadsheet whose sheet contains values associated
    with the values of `data`.

    Refer to createWorksheetFromData() for more detailed information.

    Parameters:
        data (Iterable): an iterable over values to be passed as arguments to
            the `find_value` coordinates of the Column tuples in `columns`
        columns (Iterable[Column]): an iterable over Column tuples
        sheet_title (str): the title of the active worksheet of `workbook`

    Returns:
        an openpyxl workbook object
    """
    workbook = Workbook(write_only = True)
    from openpyxl.styles import Font
    createWorksheetFromData(data, columns, workbook, title=sheet_title)
    return workbook

def createWorkbookFromGroups(groups, columns, *, sheet_prefix=None):
    """
    Create a new output spreadsheet whose sheets contains values associated
    with the values of the subiterable in each group of `groups`.

    Refer to createWorksheetsFromGroups() for more detailed information.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        sheet_prefix (Optional[str]): a title prepended onto keys when naming
            worksheets

    Returns:
        an openpyxl workbook object
    """
    workbook = Workbook(write_only=True)
    createWorksheetsFromGroups(groups, workbook, columns,
                               sheet_prefix=sheet_prefix)
    return workbook

def createWorksheetsFromGroups(groups, columns, workbook, *,
                               sheet_prefix=None):
    """
    Create new output worksheets in the spreadsheet `workbook` from the
    key-subiterable pairs in `groups`.

    Parameters:
        groups (Iterable[str : Iterable]): an iterable over tuples of keys and
            subiterables grouped by those keys (using itertools.groupby)
        columns (Iterable[Column]): an iterable over Column tuples
        workbook (openpyxl.Workbook): the workbook in which to create the sheet
        sheet_prefix (str): a title prepended onto keys when titling worksheets
    """
    for key, group in groups:
        if key is None and sheet_prefix is None:
            createWorksheetFromData(group, workbook, columns)
            continue
        elif key is not None and sheet_prefix is not None:
            title = "{} {}".format(sheet_prefix, key)
        else:
            title = str(key) if key is not None else str(sheet_prefix)
        createWorksheetFromData(group, workbook, columns, title=title)

def createWorksheetFromData(data, columns, workbook,title="Sheet1",
                            use_bold_headers=True):
    """
    Create a new output worksheet in the spreadsheet `workbook` and append
    values associated with `data` into the columns represented by `columns`.

    The values of the first row of the worksheet are obtained by the `name`
    coordinates of the Column tuples in `columns`.

    Otherwise, the value for a cell in row i and column j is obtained by
    passing the 'i - 1'th element of `data` and the integer i as arguments
    to the  `find_value` coordinate of the Column tuple in `columns` whose
    `number` coordinate is j. If no such Column exists, the cell will be empty.

    Parameters:
        data (Iterable): an iterable over values to be passed as arguments to
            the `find_value` coordinates of the Column tuples in `columns`
        columns (Iterable[Column]): an iterable over Column tuples
        workbook (openpyxl.Workbook): the workbook in which to create the sheet
        title (str): the title of the worksheet to be created
    """
    # Create the workbook and worksheet:
    worksheet = workbook.create_sheet(title=title)
    appendColumnHeadersToWorksheet(columns, worksheet, use_bold=use_bold_headers)
    appendDataToWorksheet(worksheet, data, columns)

def appendColumnHeadersToWorksheet(columns, worksheet, *, use_bold=True):
    """
    Appends the names of the Column objects in `columns` onto the excel
    worksheet `worksheet` as headers.

    Parameters:
        columns (Iterable[Column]): an iterable over Column tuples
        worksheet (openpyxl.worksheet.worksheet.Worksheet):
            the worksheet onto which headers will be appended
    """
    headers = [None] * max(column.number for column in columns)

    for column in columns:
        cell = WriteOnlyCell(worksheet, value = column.name)
        headers[column.number - 1] = cell
    worksheet.append(headers)

def appendDataToWorksheet(worksheet, data, columns, *, verbose=True):
    """
    Appends values associated with the iterable `data` onto the excel worksheet
    `worksheet`.

    Parameters:
        data (Iterable): an iterable over values to be passed as arguments to
            the `find_value` coordinates of the Column tuples in `columns`
        columns (Iterable[Column]): an iterable over Column tuples
        worksheet (openpyxl.worksheet.worksheet.Worksheet):
            the worksheet onto which data will be appended
        verbose (Optional[bool]): whether to print a message on completion
    """
    total_rows = 0
    for i, word in enumerate(data, start=worksheet._max_row):
        row = [None] * max(column.number for column in columns)
        for column in columns:
            row[column.number - 1] = column.find_value(word, i)
        worksheet.append(row)
        total_rows += 1
    if verbose: print("Wrote {} rows to {}".format(total_rows, worksheet.title))


if __name__ == '__main__':
    autoLemma(docopt(__doc__))
    exit()
